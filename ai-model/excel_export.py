from openpyxl import load_workbook
import io, os

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "template.xlsx")
MAX_ROW = 300  # must match the template's pre-filled formula rows

def generate_report(trains):
    """trains: list of dicts with trainNo, name, platform, arrival, departure, delay, trafficDensity"""
    wb = load_workbook(TEMPLATE_PATH)  # keep formulas — do NOT pass data_only=True
    raw = wb["Raw_Data"]

    # Clear only columns A-G, rows 2..MAX_ROW. Never touch H-K (formulas).
    for r in range(2, MAX_ROW + 1):
        for c in range(1, 8):
            raw.cell(row=r, column=c).value = None

    for i, t in enumerate(trains[: MAX_ROW - 1]):
        r = 2 + i
        raw.cell(row=r, column=1).value = t.get("trainNo", "")
        raw.cell(row=r, column=2).value = t.get("name", "")
        raw.cell(row=r, column=3).value = t.get("platform", "")
        raw.cell(row=r, column=4).value = t.get("arrival", "")
        raw.cell(row=r, column=5).value = t.get("departure", "")
        raw.cell(row=r, column=6).value = t.get("delay", 0)
        raw.cell(row=r, column=7).value = t.get("trafficDensity", 0)

    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
